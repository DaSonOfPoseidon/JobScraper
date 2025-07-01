# utils.py
import os
import re
import sys
import subprocess
import traceback
import asyncio
from tkinter import Tk, messagebox, simpledialog
import pandas as pd
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from dotenv import load_dotenv, set_key
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, Error as PlaywrightError
import tkinter as tk
import threading

__version__ = "0.2.1"

def get_project_root() -> str: #Returns the root directory of the project as a string path.
    if getattr(sys, "frozen", False):
        exe_path = Path(sys.executable).resolve()
        parent = exe_path.parent
        if parent.name.lower() == "bin":
            root = parent.parent
        else:
            root = parent
    else:
        file_path = Path(__file__).resolve()
        parent = file_path.parent
        if parent.name.lower() == "bin":
            root = parent.parent
        else:
            root = parent.parent  # adjust as needed
    return str(root)

# Then:
PROJECT_ROOT = get_project_root()
OUTPUT_DIR  = os.path.join(PROJECT_ROOT, "Outputs")
MISC_DIR = os.path.join(PROJECT_ROOT, "Misc")
ENV_PATH    = os.path.join(MISC_DIR, ".env")
BROWSERS    = os.path.join(PROJECT_ROOT, "browsers")
LOG_FOLDER  = os.path.join(PROJECT_ROOT, "logs")

UPDATE_MODE = None

# ensure folders exist
os.makedirs(OUTPUT_DIR, exist_ok=True)
os.makedirs(LOG_FOLDER, exist_ok=True)
os.makedirs(BROWSERS, exist_ok=True)
os.makedirs(MISC_DIR, exist_ok=True)

# === CONFIGURATION ===
load_dotenv(ENV_PATH)

BASE_URL = "http://inside.sockettelecom.com/"

class NoWOError(Exception):
    pass
class NoOpenWOError(Exception):
    pass

# Setup + Creds
def prompt_for_credentials():
    login_window = Tk()
    login_window.withdraw()

    USERNAME = simpledialog.askstring("Login", "Enter your USERNAME:", parent=login_window)
    PASSWORD = simpledialog.askstring("Login", "Enter your PASSWORD:", parent=login_window, show="*")

    login_window.destroy()
    return USERNAME, PASSWORD

def save_env_credentials(USERNAME, PASSWORD):
    dotenv_path = ENV_PATH
    if not os.path.exists(dotenv_path):
        with open(dotenv_path, "w") as f:
            f.write("")
    set_key(dotenv_path, "UNITY_USER", USERNAME)
    set_key(dotenv_path, "PASSWORD", PASSWORD)

def check_env_or_prompt_login(log=print):
    load_dotenv()
    username = os.getenv("UNITY_USER")
    password = os.getenv("PASSWORD")

    if username and password:
        log("✅ Loaded stored credentials.")
        return username, password

    while True:
        username, password = prompt_for_credentials()
        if not username or not password:
            messagebox.showerror("Login Cancelled", "Login is required to continue.")
            return None, None

        save_env_credentials(username, password)
        log("✅ Credentials captured and saved to .env.")
        return username, password

def install_chromium(log=print):
    log("=== install_chromium started ===")
    try:
        log(f"sys.frozen={getattr(sys, 'frozen', False)}")
        if getattr(sys, "frozen", False):
            log("Frozen branch: importing playwright.__main__")
            try:
                import playwright.__main__ as pw_cli
                log("Imported playwright.__main__ successfully")
            except Exception as ie:
                log(f"ImportError playwright.__main__: {ie}\n{traceback.format_exc()}")
                raise RuntimeError("Playwright package not found in the frozen bundle.") from ie

            old_argv = sys.argv.copy()
            sys.argv = ["playwright", "install", "chromium"]
            try:
                log("Calling pw_cli.main()")
                try:
                    pw_cli.main()
                    log("pw_cli.main() returned normally")
                except SystemExit as se:
                    log(f"pw_cli.main() called sys.exit({se.code}); continuing")
                    # You may check se.code: 0 means success; non-zero means failure.
                    if se.code != 0:
                        raise RuntimeError(f"playwright install exited with code {se.code}")
                except Exception as e:
                    log(f"Exception inside pw_cli.main(): {e}\n{traceback.format_exc()}")
                    raise
            finally:
                sys.argv = old_argv
        else:
            log("Script mode branch: calling subprocess")
            cmd = [sys.executable, "-m", "playwright", "install", "chromium"]
            log(f"Subprocess command: {cmd}")
            proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            log(f"Subprocess return code: {proc.returncode}")
            if proc.stdout:
                log(f"Subprocess stdout: {proc.stdout.strip()}")
            if proc.stderr:
                log(f"Subprocess stderr: {proc.stderr.strip()}")
            if proc.returncode != 0:
                raise RuntimeError(f"playwright install failed, return code {proc.returncode}")
    except Exception as e:
        log(f"Exception in install_chromium: {e}\n{traceback.format_exc()}")
        # Show error to user
        try:
            from tkinter import messagebox, Tk
            root = Tk(); root.withdraw()
            messagebox.showerror("Playwright Error", f"Failed to install Chromium:\n{e}\nSee diagnostic.log")
            root.destroy()
        except Exception as gui_e:
            print(f"Playwright install error: {e}; plus GUI error: {gui_e}")
        # Re-raise so caller knows install failed
        raise
    log("=== install_chromium finished ===")

def is_chromium_installed():
    """
    Try launching Chromium headless via sync API. Returns True if successful.
    """
    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            browser.close()
        return True
    except PlaywrightError:
        return False
    except Exception:
        return False

def ensure_playwright(log=print):
    """
    Sync check: if Chromium not installed or broken, run install_chromium().
    """
    try:
        if not is_chromium_installed():
            # Inform user
            try:
                root = Tk()
                root.withdraw()
                messagebox.showinfo("Playwright", "Chromium not found; downloading browser binaries now. This may take a few minutes.")
                root.destroy()
            except Exception:
                print("Chromium not found; downloading browser binaries now...")

            install_chromium()

            # After install, re-check
            if not is_chromium_installed():
                raise RuntimeError("Install completed but Chromium still not launchable.")
    except Exception as e:
        # Log and show error to user, referencing the log file
        err_msg = f"Playwright setup failed: {e}\nSee log file for details"
        log(err_msg)
        try:
            root = Tk()
            root.withdraw()
            messagebox.showerror("Playwright Error", err_msg)
            root.destroy()
        except Exception:
            print(err_msg)
        # Optionally exit or re-raise
        raise

def check_for_update():
    return

# Login + Session
async def handle_login(page, log=print):
    await page.goto("http://inside.sockettelecom.com/")
    # If already logged in:
    if "login.php" not in page.url:
        log("✅ Session restored with stored state.")
        await clear_first_time_overlays(page)
        return

    # Otherwise, login with creds
    user, pw = check_env_or_prompt_login(log)
    await page.goto("http://inside.sockettelecom.com/system/login.php")
    await page.fill("input[name='username']", user)
    await page.fill("input[name='password']", pw)
    await page.click("#login")
    await page.wait_for_selector("iframe#MainView", timeout=10_000)
    await clear_first_time_overlays(page)
    # Save state for next run
    await page.context.storage_state(path=os.path.join(MISC_DIR, "state.json"))
    log("✅ Logged in via credentials.")

# Browser Interaction
async def clear_first_time_overlays(page):
    selectors = [
        'input#valueForm1[type="button"][value="Close This"]',  # Vision/Mission
        'input#f1[type="button"][value="Close This"]',          # Soulkiller
    ]
    for sel in selectors:
        btn = await page.query_selector(sel)  # Non-blocking: 0ms if not present
        if btn:
            await btn.click()
            await page.wait_for_timeout(100)  # Let DOM update if needed

async def extract_cid_and_time(link, text):
    try:
        # Split lines
        lines = text.strip().split("\n")
        if len(lines) < 3:
            return None, None, None
        raw_time = lines[0].strip()
        # If it's a range, keep only the first hour/block
        if "-" in raw_time:
            first = raw_time.split("-", 1)[0].strip()
            # Optionally normalize: if it's just "1", you could append ":00" -> "1:00"
            # For now we keep as-is:
            time_slot = first
        else:
            time_slot = raw_time
        # lines[1] = job type/area, usually not needed here
        third_line = lines[2].strip()
        # Now split the third line by ' - '
        parts = third_line.split(" - ")
        if len(parts) < 3:
            return None, None, None
        name = parts[0].strip()
        cid = parts[1].strip()
        # Optional: parse order #
        order_match = re.search(r'Order\s*#:(\d+)', parts[2])
        order_num = order_match.group(1) if order_match else None
        return cid, name, time_slot
    except Exception as e:
        print(f"❌ Error extracting CID/time: {e}")
        return None, None, None

async def get_contractor_assignments(page):
    try:
        # Wait for the contractor section (parent) and ContractorList (child) to be visible
        await page.wait_for_selector(".contractorsection #ContractorList", timeout=15_000, state="visible")
        # Now, get all <b> elements inside the contractor list
        contractor_b_tags = await page.locator(".contractorsection #ContractorList b").all_inner_texts()
        for btext in contractor_b_tags:
            if "None Assigned" in btext:
                return "None Assigned"
            if " - (Primary" in btext:
                return btext.split(" - ")[0].strip()
            elif btext.strip() and "assigned to this work order" not in btext:
                # fallback for any other contractor <b>
                return btext.strip()
        return "Unknown"
    except Exception as e:
        print(f"❌ Could not extract contractor: {e}")
        return "Unknown"

async def get_work_order_url(frame, log=print):
    """
    Find newest (highest-numbered) in-process Fiber Install WO on customer page.
    Raises NoWOError or NoOpenWOError if not found.
    Returns (absolute_url, wo_number).
    """
    try:
        # Wait for the work order table to load (more general selector)
        try:
            await frame.wait_for_selector("#custWork #workShow table", timeout=10_000)
        except Exception:
            log("❌ Work Orders table not found inside frame!")
            raise NoWOError("Work Orders table not found!")
        rows = await frame.query_selector_all("#custWork #workShow table tr")
        if not rows:
            log("❌ No work order rows found inside table.")
            raise NoWOError("No work order rows found!")

        matches = []
        for row in rows:
            tds = await row.query_selector_all("td")
            if len(tds) < 5:
                continue

            first = (await tds[0].inner_text()).strip()
            # skip header or invalid rows
            if first == "#" or not first.isdigit():
                continue

            wo_number = int(first)
            job_type = (await tds[2].inner_text()).strip()
            status   = (await tds[3].inner_text()).strip()
            # Only match "Fiber Install" AND "In Process"
            if all(s in job_type for s in ["Fiber", "Install"]) and "In Process" in status:
                link_td = await tds[4].query_selector("a")
                href = await link_td.get_attribute('href') if link_td else None
                matches.append((wo_number, href))

        if not matches:
            # Fallback: any "Fiber Install", even if not "In Process"
            for row in rows:
                tds = await row.query_selector_all("td")
                if len(tds) < 5:
                    continue
                job_type = (await tds[2].inner_text()).strip()
                if all(s in job_type for s in ["Fiber", "Install"]):
                    raise NoOpenWOError("No open (In Process) Fiber Install WO found.")
            raise NoWOError("No Fiber Install WO found at all.")

        # Return WO with the highest number (most recent)
        wo_number, url = max(matches, key=lambda x: x[0])
        # Make sure URL is absolute
        if url and url.startswith("/"):
            url = "http://inside.sockettelecom.com" + url
        return url, wo_number

    except NoWOError:
        raise
    except NoOpenWOError:
        raise
    except Exception as e:
        log(f"❌ Error in get_work_order_url: {e}")
        raise NoWOError(str(e))

async def get_job_type_and_address(page):
    # 1) Grab the address for city-inspection
    address = "Unknown"
    try:
        addr_elem = await page.query_selector("a[href*='viewServiceMap']")
        if addr_elem:
            address = (await addr_elem.inner_text()).strip()
    except Exception:
        pass

    # 2) Read the packageName <b> text for actual package info
    package_info = ""
    try:
        pkg_elem = await page.query_selector(".packageName.text-indent b")
        if pkg_elem:
            package_info = (await pkg_elem.inner_text()).strip()
    except Exception:
        pass
    pkg_lower = package_info.lower()

    # 3) Grab the description (for connectorized check)
    desc_text = ""
    try:
        desc_elem = await page.query_selector("td.detailHeader:has-text('Description:') + td.detailData")
        if desc_elem:
            desc_text = (await desc_elem.inner_text()).strip().lower()
        else:
            # fallback for XPath
            desc_elem = await page.query_selector("//td[contains(text(), 'Description:')]/following-sibling::td[1]")
            if desc_elem:
                desc_text = (await desc_elem.inner_text()).strip().lower()
    except Exception:
        pass

    # 4) OFFICIAL 5 Gig check 
    if "5 gig" in pkg_lower and "2.5" not in pkg_lower:
        is_connectorized = "connectorized" in desc_text
        has_phone        = "bundle" in pkg_lower or "phone" in pkg_lower

        if is_connectorized:
            job_type = "Connectorized 5 Gig Bundle" if has_phone else "Connectorized 5 Gig"
        else:
            job_type = "5 Gig Fiber Bundle"       if has_phone else "5 Gig Naked Fiber"

        return job_type, address

    # 5) Conversion Fallback: no package + Jefferson City → 5 Gig Conversion (until IT fixes the WO)
    if not package_info and "jefferson city" in address.lower():
        return "5 Gig Conversion", address

    # 6) 2.5G branch
    if "2.5" in pkg_lower:
        is_connectorized = "connectorized" in desc_text
        has_phone        = "bundle" in pkg_lower or "phone" in pkg_lower

        if is_connectorized:
            job_type = "Connectorized 2.5G Bundle" if has_phone else "Connectorized 2.5G"
        else:
            job_type = "2.5G Fiber Bundle"       if has_phone else "2.5G Naked Fiber"

        return job_type, address

    # 7) OTHERWISE: your existing Connectorized / Bundle / Naked logic
    is_connectorized = "connectorized" in desc_text
    has_phone        = "bundle" in pkg_lower or "phone" in pkg_lower

    if is_connectorized:
        job_type = "Connectorized Bundle" if has_phone else "Connectorized"
    else:
        job_type = "Fiber Bundle"       if has_phone else "Naked Fiber"

    return job_type, address

async def extract_wo_date(page, fallback_date=None):
    try:
        # Wait for the scheduled event section to load
        await page.wait_for_selector("#scheduledEventList", timeout=10_000)

        # Now poll until the date string appears or timeout reached
        deadline = asyncio.get_event_loop().time() + 8  # up to 8 extra seconds
        date_pattern = re.compile(r"\d{4}-\d{2}-\d{2}")

        while True:
            text = (await page.locator("#scheduledEventList").inner_text()).strip()
            if date_pattern.search(text):
                break
            # Also break if the content is NOT a spinner/loader, e.g., contains "Fiber Install"
            if "Fiber" in text and "Install" in text:
                break
            if asyncio.get_event_loop().time() > deadline:
                break
            await asyncio.sleep(0.25)

        # Continue as before
        text = re.sub(r"<.*?>", "", text)
        lines = text.splitlines()
        fiber_line = next(
            (line for line in lines if re.search(r"fiber.*install", line, re.I)), ""
        )
        date_match = re.search(r"(\d{4})-(\d{2})-(\d{2})", fiber_line)
        if date_match:
            parsed_date = datetime.strptime(date_match.group(0), "%Y-%m-%d")
            fmt = "%#m-%#d-%y" if os.name == "nt" else "%-m-%-d-%y"
            return parsed_date.strftime(fmt)

        # Fallbacks (as above)
        alt_match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", fiber_line)
        if alt_match:
            try:
                mm, dd, yy = map(int, alt_match.groups())
                if yy < 100:  # Two-digit year
                    yy += 2000
                parsed_date = datetime(yy, mm, dd)
                fmt = "%#m-%#d-%y" if os.name == "nt" else "%-m-%-d-%y"
                return parsed_date.strftime(fmt)
            except Exception:
                pass

        if fallback_date:
            return fallback_date

    except Exception as e:
        print(f"⚠️ Could not extract WO date: {e}")
    return "Unknown"

async def assign_contractor(page, wo_number, desired_contractor_full, log=print):
    try:
        await page.wait_for_load_state("domcontentloaded")
        await page.wait_for_selector("#ContractorList", state="visible", timeout=10000)
        assigned_contractors = []
        contractor_rows = page.locator("#ContractorList table tbody tr")
        count = await contractor_rows.count()
        for i in range(count):
            row = contractor_rows.nth(i)
            contractor_name = await row.locator("td b").inner_text()
            assigned_contractors.append(contractor_name.strip())

        # Exact match (case insensitive)
        if any(c.lower() == desired_contractor_full.lower() for c in assigned_contractors):
            log(f"✅ Contractor '{desired_contractor_full}' already assigned to WO #{wo_number}")
        else:
            assign_link = page.locator("b.addattachlink", has_text="Assign Contractor(s)")
            await assign_link.click()

            await page.wait_for_function(
                """() => {
                    const el = document.querySelector('#ContractorAddArea');
                    return el && window.getComputedStyle(el).display === 'none';
                }""",
                timeout=5000
            )

            contractor_dropdown = page.locator("#ContractorID")
            await contractor_dropdown.select_option(label=desired_contractor_full)
            role_dropdown = page.locator("#ContractorType")
            await role_dropdown.select_option(label="Primary")

            assign_button = page.locator("input[type='button'][value='Assign']")
            await assign_button.wait_for(state="visible", timeout=5000)
            await assign_button.scroll_into_view_if_needed()
            await assign_button.click()
            await page.wait_for_timeout(300)  # Allow UI to update

            await page.wait_for_selector("#ContractorAddArea", state="hidden", timeout=5000)

            #log(f"🏷️ Assigned contractor '{desired_contractor_full}' to WO #{wo_number}")

            # 5) Remove any other contractors (other than desired one)
            while True:
                contractor_rows = page.locator("#ContractorList table tbody tr")
                count = await contractor_rows.count()
                removed_any = False
                for i in range(count):
                    row = contractor_rows.nth(i)
                    contractor_name = await row.locator("td b").inner_text()
                    contractor_name = contractor_name.strip()
                    if desired_contractor_full not in contractor_name:
                        #log(f"🔍 Removing contractor: '{contractor_name}'")
                        remove_link = row.locator("td a", has_text="Remove")
                        try:
                            await remove_link.click()
                            await page.wait_for_timeout(300)  # Wait for DOM update
                            removed_any = True
                            break  # Refresh the list after DOM changes
                        except Exception as e:
                            log(f"❌ Failed to remove contractor '{contractor_name}'")
                if not removed_any:
                    break

    except Exception as e:
        log(f"❌ Contractor assignment process failed for WO #{wo_number}")

def prompt_reassignment(root, spread_file, log_func=print):
    """
    Pops up a modal dialog asking to apply contractor reassignments.
    If user agrees, runs the reassignment asynchronously on a background thread.
    """
    def start_reassignment():
        popup.destroy()
        threading.Thread(target=lambda: asyncio.run(apply_spread_changes(spread_file, log_func)), daemon=True).start()

    def cancel():
        popup.destroy()

    popup = tk.Toplevel(root)
    popup.title("Apply Spread Changes?")
    tk.Label(popup, text="Apply contractor reassignments now?").pack(padx=20, pady=10)

    btn_frame = tk.Frame(popup)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="Reassign", command=start_reassignment, width=12).pack(side="left", padx=10)
    tk.Button(btn_frame, text="Not Now", command=cancel, width=12).pack(side="right", padx=10)

    popup.grab_set()
    popup.transient(root)
    popup.wait_window()

async def apply_spread_changes(spread_file, log_func=print):
    from scrape_runner import init_playwright_page
    from spreader import parse_moved_jobs_from_spread
    jobs = parse_moved_jobs_from_spread(spread_file)
    if not jobs:
        log_func("No moved jobs to reassign.")
        return
    playwright, browser, context, page = await init_playwright_page(headless=True)
    try:
        await handle_login(page, log=log_func)
        for job in jobs:
            wo_number = job["wo"]
            desired_contractor = job["contractor"]
            try:
                url = f"http://inside.sockettelecom.com/workorders/view.php?nCount={wo_number}"
                await page.goto(url)
                await asyncio.sleep(1)  # let page settle
                await assign_contractor(page, wo_number, desired_contractor, log=log_func)
            except Exception as e:
                log_func(f"Failed to process WO {wo_number}: {e}")
        log_func("Done applying spread changes.")
    finally:
        await page.close()
        await context.close()
        await browser.close()
        await playwright.stop()

# Time + Data
def get_output_tag(start, end): #File date stamp
    if start == end:
        return start.strftime("%m%d")
    return f"{start.strftime('%m%d')}-{end.strftime('%m%d')}"

def parse_time(t):
    try:
        return datetime.strptime(t.strip(), "%I:%M")
    except:
        return datetime.min

def get_sort_key(time_str):
    time_str = time_str.strip()
    hour = int(time_str.split(":")[0])
    # Treat hours before 6 as PM
    if hour in [1, 2, 3, 4, 5]:
        hour += 12
    return hour

def parse_date(date_str):
    try:
        return datetime.strptime(date_str.strip(), "%m-%d-%y").date()
    except ValueError as e:
        print(f"Warning: failed to parse date '{date_str}': {e}")
        return datetime.min.date()

def company_sort_key(name):
    if name == "Unknown":
        return (0, "")          # first
    elif name == "None Assigned":
        return (2, "")          # last
    else:
        return (1, name.lower()) # middle, alphabetically

# I/O
def generate_changes_file(old_list, new_list, changes_filename):
    def stringify(j):
        return f"{j['time']} - {j['name']} - {j['cid']} - {j['type']} - {j['address']} - WO {j['wo']}"

    # build sets of lines per company
    old_by_co = defaultdict(set)
    new_by_co = defaultdict(set)
    for j in old_list:
        old_by_co[j['company']].add(stringify(j))
    for j in new_list:
        new_by_co[j['company']].add(stringify(j))

    # all companies seen
    companies = sorted(set(old_by_co) | set(new_by_co))

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    path = os.path.join(OUTPUT_DIR, changes_filename)

    with open(path, 'w', encoding='utf-8') as f:
        for co in companies:
            f.write(f"{co}\n")
            f.write("Added:\n")
            for line in sorted(new_by_co[co] - old_by_co.get(co, set())):
                f.write(f"  {line}\n")
            f.write("\nRemoved:\n")
            for line in sorted(old_by_co[co] - new_by_co.get(co, set())):
                f.write(f"  {line}\n")
            f.write("\n")

    return path

def export_txt(jobs, filename=None):
    jobs_by_company = defaultdict(lambda: defaultdict(list))
    for job in jobs:
        jobs_by_company[job["company"]][job["date"]].append(job)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_name = os.path.basename(filename) if filename else "Jobs.txt"
    out_path = os.path.join(OUTPUT_DIR, out_name)

    with open(out_path, "w", encoding="utf-8") as f:
        for company in sorted(jobs_by_company.keys(), key=company_sort_key):
            f.write(f"{company}\n\n")
            days = jobs_by_company[company]

            # Sort dates chronologically by parsing date strings
            sorted_dates = sorted(days.keys(), key=parse_date)

            for date in sorted_dates:
                f.write(f"{date}\n")
                entries = days[date]

                # Sort entries by time, then by customer name
                entries_sorted = sorted(entries, key=lambda j: (get_sort_key(j['time']), j['name'].lower()))
                for job in entries_sorted:
                    f.write(f"{job['time']} - {job['name']} - {job['cid']} - {job['type']} - {job['address']} - WO {job['wo']}\n")
                f.write("\n")
            f.write("\n")

def export_excel(jobs, filename=None):
    jobs_by_company = defaultdict(lambda: defaultdict(list))
    for job in jobs:
        jobs_by_company[job["company"]][job["date"]].append(job)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_name = os.path.basename(filename) if filename else "Jobs.txt"
    out_path = os.path.join(OUTPUT_DIR, out_name)

    rows = []
    for company, days in jobs_by_company.items():
        rows.append([company])
        for date, entries in sorted(days.items()):
            rows.append([date])
            for job in sorted(entries, key=lambda j: (get_sort_key(j['time']), j['name'].lower())):
                rows.append([
                    job['time'],
                    job['name'],
                    job['cid'],
                    job['type'],
                    job['address'],
                    f"WO {job['wo']}"
                ])
            rows.append([])
        rows.append([])

    # Write to Excel
    df = pd.DataFrame(rows)
    
    df.to_excel(out_path, index=False, header=False)

    # Autofit column widths
    wb = load_workbook(out_path)
    ws = wb.active
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(out_path)

def parse_imported_jobs(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    jobs = []

    if ext == ".txt":
        with open(file_path, "r", encoding="utf-8") as f:
            current_company = None
            current_date    = None

            for raw in f:
                line = raw.strip()
                if not line:
                    continue

                # 1) Company header: any line without " - " and no digits
                if " - " not in line and not any(c.isdigit() for c in line):
                    current_company = line
                    continue

                # 2) Date header: matches M-D-YY or M-D-YYYY
                if re.match(r'^\d{1,2}-\d{1,2}-\d{2,4}$', line):
                    current_date = line
                    continue

                # 3) Job line
                if " - " in line and "WO" in line:
                    parts = [p.strip() for p in line.split(" - ")]
                    if len(parts) >= 6:
                        time, name, cid, typ, addr, wo = parts[:6]
                        jobs.append({
                            "company": current_company,
                            "date":    current_date,
                            "time":    time,
                            "name":    name,
                            "cid":     cid,
                            "type":    typ,
                            "address": addr,
                            "wo":      wo.replace("WO ", "")
                        })

    elif ext == ".xlsx":
        df = pd.read_excel(file_path, header=None)
        current_company = None
        current_date    = None

        for row in df.itertuples(index=False):
            # flatten row to single string list
            cells = [str(c).strip() for c in row if c and str(c).strip()]
            if not cells:
                continue

            line = " - ".join(cells)
            # same detection logic as above
            if " - " not in line and not any(c.isdigit() for c in line):
                current_company = line
                continue
            if re.match(r'^\d{1,2}-\d{1,2}-\d{2,4}$', line):
                current_date = line
                continue
            if " - " in line and "WO" in line:
                parts = [p.strip() for p in line.split(" - ")]
                if len(parts) >= 6:
                    time, name, cid, typ, addr, wo = parts[:6]
                    jobs.append({
                        "company": current_company,
                        "date":    current_date,
                        "time":    time,
                        "name":    name,
                        "cid":     cid,
                        "type":    typ,
                        "address": addr,
                        "wo":      wo.replace("WO ", "")
                    })

    return jobs
