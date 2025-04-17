import os
import re
import time
import traceback
import pandas as pd
from datetime import datetime
from collections import defaultdict
from dotenv import load_dotenv
load_dotenv(dotenv_path=".env")

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException

import tkinter as tk
from tkinter import filedialog, messagebox, Label, Button
from tkinterdnd2 import DND_FILES, TkinterDnD

# === Constants ===
CALENDAR_URL = "http://inside.sockettelecom.com/events/calendar.php"
CUSTOMER_URL_TEMPLATE = "http://inside.sockettelecom.com/menu.php?coid=1&tabid=7&parentid=9&customerid={}"

HEADLESS = os.getenv("HEADLESS", "False").lower() == "true"
USERNAME = os.getenv("UNITY_USER")
PASSWORD = os.getenv("PASSWORD")

if not USERNAME or not PASSWORD:
    raise RuntimeError("Missing UNITY_USER or PASSWORD in .env file.")

# === Utility Functions ===
def dismiss_alert(driver):
    try:
        for _ in range(3):  # Try up to 3 alerts
            WebDriverWait(driver, 1).until(EC.alert_is_present())
            alert = driver.switch_to.alert
            #print(f"⚠️ Dismissing alert: {alert.text}")
            alert.dismiss()
            time.sleep(0.5)
    except:
        pass

def force_dismiss_any_alert(driver):
    try:
        WebDriverWait(driver, 1).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        #print(f"⚠️ Force dismissing alert: {alert.text}")
        alert.dismiss()
        time.sleep(0.25)
    except:
        pass

def clear_first_time_overlays(driver):
    try:
        WebDriverWait(driver, 2).until(EC.alert_is_present())
        alert = driver.switch_to.alert
        #print(f"⚠️ Dismissing alert: {alert.text}")
        alert.dismiss()
        time.sleep(0.25)
    except:
        pass

    # Try to close Vision modal fast
    try:
        close_btn = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.XPATH, "//form[@id='valueForm']//input[@type='button']"))
        )
        #print("⚠️ Closing Vision modal...")
        close_btn.click()
        time.sleep(0.25)
    except:
        pass

    # Try to close soulkiller popup fast
    try:
        soulkiller_close = WebDriverWait(driver, 2).until(
            EC.element_to_be_clickable((By.XPATH, "//form[@id='f']//input[@type='button']"))
        )
        print("Closing soulkiller popup...")
        soulkiller_close.click()
        time.sleep(0.25)
    except:
        pass

    # Try switching to iframe quickly with fallback loop
    for _ in range(4):  # Retry up to ~4x within ~2s total
        try:
            WebDriverWait(driver, 0.5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "MainView")))
            #print("✅ Switched into MainView iframe.")
            return
        except:
            time.sleep(0.5)
    print("❌ Could not switch to MainView iframe.")

def extract_cid_and_time(job_elem):
    try:
        text = job_elem.find_element(By.CLASS_NAME, "fc-title").text
        time_text = job_elem.find_element(By.CLASS_NAME, "fc-time").get_attribute("data-start")
        name, cid = re.findall(r"(.*?) - (\d{4}-\d{4}-\d{4})", text)[0]
        return cid.strip(), name.strip(), time_text.strip()
    except Exception as e:
        print(f"\u274c Error extracting CID/time: {e}")
        return None, None, None

def get_contractor_assignments(driver):
    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, "contractorsection"))
        )
        section_elem = driver.find_element(By.CLASS_NAME, "contractorsection")
        section_text = section_elem.text.strip()
        # print(f"[DEBUG] contractorsection raw text:\n{section_text}")

        # Extract the contractor name (line containing 'Primary')
        lines = section_text.splitlines()
        for line in lines:
            if " - (Primary" in line:
                contractor_name = line.split(" - ")[0].strip()
                #print(f"[DEBUG] contractor_name parsed: {contractor_name}")
                return contractor_name

        # Fallback if 'Primary' not found, just use first non-header line
        for line in lines:
            if "assigned to this work order" not in line and "Contractors" not in line:
                contractor_name = line.split(" - ")[0].strip()
                #print(f"[DEBUG] contractor_name fallback: {contractor_name}")
                return contractor_name

        #print("[DEBUG] No contractor lines matched.")
        return "Unknown"

    except Exception as e:
        #print("[DEBUG] contractorsection not found or failed to load.")
        print(f"❌ Could not find contractor name: {e}")
        return "Unknown"

def get_work_order_url(driver):
    try:
        dismiss_alert(driver)

        # Faster polling
        wait = WebDriverWait(driver, 5, poll_frequency=0.05)
        wait.until(
            lambda d: d.find_element(By.ID, "workShow").is_displayed()
        )

        workorder_rows = driver.find_elements(By.XPATH, "//div[@id='workShow']//table//tr[position()>1]")

        if not workorder_rows:
            print("⚠️ Customer has no work orders.")
            return None, None

        install_wos = []
        for row in workorder_rows:
            cols = row.find_elements(By.TAG_NAME, "td")
            wo_num = cols[0].text.strip()
            desc = cols[1].text.strip()
            status = cols[3].text.strip()
            url = cols[4].find_element(By.TAG_NAME, "a").get_attribute("href")

            desc_lower = desc.lower()
            if "install" in desc_lower and "fiber" in desc_lower:
                install_wos.append((int(wo_num), status, url))

        for wo in install_wos:
            if wo[1].lower() == "in process":
                return wo[2], wo[0]

        if install_wos:
            max_wo = max(install_wos, key=lambda x: x[0])
            return max_wo[2], max_wo[0]

        print("⚠️ No Fiber Install WOs found, even though customer has work orders.")
        return None, None

    except Exception as e:
        print(f"❌ Could not find work order URL: {e}")
        return None, None

def get_job_type_and_address(driver):
    wait = WebDriverWait(driver, 10)
    job_type = "Unknown"
    address = "Unknown"

    is_connectorized = False
    has_phone = False

    # === Step 1: Check Work Order Description for Legacy vs Connectorized ===
    try:
        desc_elem = driver.find_element(By.XPATH, "//td[contains(text(), 'Description:')]/following-sibling::td")
        description_text = desc_elem.text.strip().lower()
        #print(f"[DEBUG] WO Description: {description_text}")
        is_connectorized = "connectorized" in description_text
    except Exception as e:
        print(f"❌ Failed to get WO description for connectorized check: {e}")

    # === Step 2: Check Services Section for Phone/Bundles ===
    try:
        wait.until(lambda d: d.find_element(By.CLASS_NAME, "servicesDiv").text.strip() != "")
        service_div = driver.find_element(By.CLASS_NAME, "servicesDiv").text.strip().lower()
        #print(f"[DEBUG] servicesDiv: {service_div}")
        has_phone = "phone" in service_div # add any other phone indicators here
    except Exception as e:
        print(f"❌ Failed to check service info for phone bundle: {e}")

    # === Final Job Type Decision ===
    if is_connectorized:
        job_type = "Connectorized Bundle" if has_phone else "Connectorized"
    else:
        job_type = "Fiber Bundle" if has_phone else "Naked Fiber"

    # === Address Detection ===
    try:
        address_elem = driver.find_element(By.XPATH, "//a[contains(@href, 'viewServiceMap')]")
        address = address_elem.text.strip()
        #print(f"[DEBUG] address found: {address}")
    except Exception as e:
        print(f"❌ get_job_type_and_address() - unexpected error getting address: {e}")

    #print(f"[DEBUG] Final job type: {job_type}")
    return job_type, address

def extract_wo_date(driver):
    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "scheduledEventList"))
        )
        container = driver.find_element(By.ID, "scheduledEventList")
        text = container.text.strip()
        #print(f"[DEBUG] scheduledEventList text:\n{text}")

        # Split and scan only relevant install entries
        lines = text.splitlines()
        install_line = next((line for line in lines if "Fiber Install" in line), None)

        if not install_line:
            print("⚠️ No Residential Fiber Install event found in scheduledEventList.")
            return "Unknown"

        match = re.search(r"(\d{4}-\d{2}-\d{2})", install_line)
        if match:
            date_str = match.group(1)
            parsed_date = datetime.strptime(date_str, "%Y-%m-%d")
            formatted = parsed_date.strftime("%#m-%#d-%y") if os.name == "nt" else parsed_date.strftime("%-m-%-d-%y")
            #print(f"[DEBUG] parsed WO date: {formatted}")
            return formatted

        print("⚠️ Date not found in install line.")
        return "Unknown"

    except Exception as e:
        print(f"⚠️ Could not extract WO date: {e}")
        return "Unknown"

def scrape_jobs(driver, mode="new", imported_jobs=None, selected_day=None, test_mode=False, test_limit=10):
    driver.get(CALENDAR_URL)
    wait = WebDriverWait(driver, 30)
    main_window = driver.current_window_handle

    force_dismiss_any_alert(driver)

    try:
        force_dismiss_any_alert(driver)
        week_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'week')]")))
        week_button.click()
        time.sleep(1)
        wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))
        print("✅ Switched to Week View.")
    except Exception as e:
        print(f"⚠️ Could not switch to Week View: {e}")

    try:
        force_dismiss_any_alert(driver)
        next_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "button.fc-next-button")))
        next_button.click()
        time.sleep(1)
        wait.until(EC.invisibility_of_element_located((By.ID, "spinner")))
        print("✅ Navigated to Next Week.")
    except Exception as e:
        print(f"⚠️ Could not advance to next week: {e}")

    try:
        wait.until(lambda d: any("Residential Fiber Install" in el.text for el in d.find_elements(By.CSS_SELECTOR, "a.fc-time-grid-event")))
        print("✅ Jobs loaded and ready to scrape.")
    except:
        print("⚠️ No 'Residential Fiber Install' jobs detected.")

    job_links = driver.find_elements(By.CSS_SELECTOR, "a.fc-time-grid-event")
    results = []
    customer_counter = 0

    try:
        for link in job_links:
            if "Residential Fiber Install" not in link.text:
                continue

            cid, name, time_slot = extract_cid_and_time(link)
            if not cid:
                continue

            customer_counter += 1
            timestamp = datetime.now().strftime("[%H:%M:%S]")
            print(f"{timestamp} {customer_counter} - {cid} - Opening")

            try:
                customer_url = CUSTOMER_URL_TEMPLATE.format(cid)
                driver.execute_script("window.open('');")
                driver.switch_to.window(driver.window_handles[-1])
                driver.get(customer_url)
                clear_first_time_overlays(driver)

                workorder_url, wo_number = get_work_order_url(driver)
                if not workorder_url:
                    print(f"⚠️ No WO found for {cid}")
                    driver.close()
                    driver.switch_to.window(main_window)
                    continue

                try:
                    driver.get(workorder_url)
                except Exception as we:
                    print(f"❌ Timeout or error loading WO page for CID {cid}: {we}")
                    driver.close()
                    driver.switch_to.window(main_window)
                    continue

                dismiss_alert(driver)
                time.sleep(1)

                job_type, address = get_job_type_and_address(driver)

                if mode == "update" and imported_jobs:
                    match = next((j for j in imported_jobs if j["cid"] == cid and j["address"] == address), None)
                    if match:
                        print(f"🔁 Skipping existing job: {cid} at {address}")
                        driver.close()
                        driver.switch_to.window(main_window)
                        continue

                contractor_info = get_contractor_assignments(driver)
                job_date = extract_wo_date(driver)

                results.append({
                    "company": contractor_info,
                    "date": job_date,
                    "time": time_slot,
                    "name": name,
                    "cid": cid,
                    "type": job_type,
                    "address": address,
                    "wo": wo_number
                })

            except Exception as e:
                print(f"❌ Error while processing job for CID {cid}: {e}")
                traceback.print_exc()

            finally:
                try:
                    if len(driver.window_handles) > 1:
                        driver.close()
                    driver.switch_to.window(main_window)
                except Exception as e:
                    print(f"❌ Failed to switch back to main window: {e}")
                    break

            if test_mode and len(results) >= test_limit:
                print("🔬 Test mode: Job limit reached. Exiting early.")
                break

    finally:
        if test_mode:
            print(f"🔬 Test mode: Exporting {len(results)} test results...")
            export_txt(results, "calendar_test_output.txt")
            export_excel(results, "calendar_test_output.xlsx")
        elif results:
            print(f"⚠️ Exporting {len(results)} partial results before exit...")
            export_txt(results, "calendar_partial_output.txt")
            export_excel(results, "calendar_partial_output.xlsx")
        else:
            print("❌ No results to export.")

    print(f"✅ Scraped {len(results)} jobs.")
    return results


# === GUI ===
class CalendarBuddyGUI:
    def __init__(self, root):
        self.root = root
        self.driver = None
        self.root.geometry("520x260")
        self.root.configure(bg="#f0f0f0")

        # File drop label
        self.label = Label(root, text="Drag and drop a .txt or .xlsx file here", width=50, height=4, bg="white", relief="ridge")
        self.label.grid(row=0, column=0, columnspan=3, padx=10, pady=10)

        # Test Mode checkbox
        self.test_mode = tk.BooleanVar()
        self.test_checkbox = tk.Checkbutton(root, text="Test Mode", variable=self.test_mode)
        self.test_checkbox.grid(row=1, column=0, sticky="w", padx=10)

        # Test Mode limit spinbox
        self.limit_label = Label(root, text="Test Mode Limit:")
        self.limit_label.grid(row=1, column=1, sticky="e")
        self.test_limit = tk.IntVar(value=10)
        self.limit_spinbox = tk.Spinbox(root, from_=1, to=500, textvariable=self.test_limit, width=5)
        self.limit_spinbox.grid(row=1, column=2, sticky="w")

        # Headless Mode toggle
        self.headless_mode = tk.BooleanVar(value=HEADLESS)
        self.headless_checkbox = tk.Checkbutton(root, text="Headless Mode", variable=self.headless_mode)
        self.headless_checkbox.grid(row=2, column=0, sticky="w", padx=10)

        # Browse button
        self.open_button = Button(root, text="Or Click to Browse", command=self.browse_file)
        self.open_button.grid(row=2, column=2, sticky="e", padx=10)

        # Action buttons
        self.scrape_button = Button(root, text="Get Weekly Jobs", command=self.get_weekly_jobs)
        self.scrape_button.grid(row=3, column=0, columnspan=3, pady=10)

        self.update_button = Button(root, text="Update Jobs", command=self.update_jobs)
        self.update_button.grid(row=4, column=0, columnspan=3)

        # Drop handler
        self.label.drop_target_register(DND_FILES)
        self.label.dnd_bind('<<Drop>>', self.handle_drop)


    def handle_drop(self, event):
        self.dropped_file_path = event.data.strip('{}')
        self.process_file()

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("Excel files", "*.xlsx")])
        if file_path:
            self.dropped_file_path = file_path
            self.process_file()

    def process_file(self):
        if self.dropped_file_path:
            ext = os.path.splitext(self.dropped_file_path)[1].lower()
            if ext == ".txt":
                print(f"📝 Text file loaded: {self.dropped_file_path}")
            elif ext == ".xlsx":
                print(f"📊 Excel file loaded: {self.dropped_file_path}")
            else:
                print("❌ Unsupported file type.")
        else:
            print("⚠️ No file selected.")

    def get_weekly_jobs(self):
        try:
            self.init_driver()
            if not self.driver:
                return  # Stop here if driver init failed
            jobs = scrape_jobs(self.driver, mode="new", test_mode=self.test_mode.get(), test_limit=self.test_limit.get())
            if not jobs:
                messagebox.showwarning("No Jobs", "No new jobs were found.")
                return
            export_txt(jobs)
            export_excel(jobs)
            messagebox.showinfo("Export Complete", f"Exported {len(jobs)} jobs to .txt and .xlsx files.")
        except Exception as e:
            messagebox.showerror("Error", f"Something went wrong: {e}")

    def update_jobs(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt"), ("Excel files", "*.xlsx")])
        if not file_path:
            return
        imported_jobs = self.load_jobs_from_file(file_path)
        if imported_jobs is None:
            messagebox.showerror("Error", "Could not load jobs from the file.")
            return
        try:
            self.init_driver()
            if not self.driver:
                return  # Stop here if driver init failed
            jobs = scrape_jobs(self.driver, mode="update", imported_jobs=imported_jobs, test_mode=self.test_mode.get(), test_limit=self.test_limit.get())

            if not jobs:
                messagebox.showinfo("No New Jobs", "No new jobs found to add.")
                return
            export_txt(jobs, "calendar_updated.txt")
            export_excel(jobs, "calendar_updated.xlsx")
            messagebox.showinfo("Update Complete", f"Exported {len(jobs)} new jobs to calendar_updated.txt and .xlsx.")
        except Exception as e:
            messagebox.showerror("Scraping Error", f"Something went wrong while updating: {e}")

    def load_jobs_from_file(self, file_path):
        jobs = []
        ext = os.path.splitext(file_path)[1].lower()
        try:
            if ext == ".txt":
                with open(file_path, "r") as f:
                    for line in f:
                        if "-" in line and "WO" in line:
                            parts = line.strip().split(" - ")
                            if len(parts) >= 6:
                                jobs.append({
                                    "time": parts[0],
                                    "name": parts[1],
                                    "cid": parts[2],
                                    "type": parts[3],
                                    "address": parts[4],
                                    "wo": parts[5].replace("WO ", "").strip()
                                })
            elif ext == ".xlsx":
                df = pd.read_excel(file_path, header=None)
                for row in df.itertuples(index=False):
                    if len(row) >= 6 and isinstance(row[5], str) and "WO" in row[5]:
                        jobs.append({
                            "time": str(row[0]),
                            "name": str(row[1]),
                            "cid": str(row[2]),
                            "type": str(row[3]),
                            "address": str(row[4]),
                            "wo": row[5].replace("WO ", "").strip()
                        })
        except Exception as e:
            print(f"❌ Failed to parse imported file: {e}")
            return None
        return jobs

    def init_driver(self):
        try:
            if self.driver is None:
                options = Options()
                options.page_load_strategy = 'eager'
                options.add_experimental_option("detach", True)

                if self.headless_mode.get():
                    options.add_argument("--headless=new")
                    options.add_argument("--disable-gpu")
                    options.add_argument("--window-size=1920,1080")
                    options.add_argument("--no-sandbox")
                    options.add_argument("--disable-dev-shm-usage")   

                self.driver = webdriver.Chrome(service=Service("./chromedriver.exe"), options=options)
                handle_login(self.driver)
        except Exception as e:
            messagebox.showerror("Login Error", f"Login failed: {e}")
            self.driver = None


# === EXPORT FUNCTIONS ===
def export_txt(jobs, filename="calendar_output.txt"):
    jobs_by_company = defaultdict(lambda: defaultdict(list))
    for job in jobs:
        jobs_by_company[job["company"]][job["date"]].append(job)
    with open(filename, "w") as f:
        for company, days in jobs_by_company.items():
            f.write(f"{company}\n\n")
            for date, entries in sorted(days.items()):
                f.write(f"{date}\n")
                for job in entries:
                    line = f"{job['time']} - {job['name']} - {job['cid']} - {job['type']} - {job['address']} - WO {job['wo']}"
                    f.write(line + "\n")
                f.write("\n")
            f.write("\n")

def export_excel(jobs, filename="calendar_output.xlsx"):
    from openpyxl.utils import get_column_letter
    from openpyxl import load_workbook

    jobs_by_company = defaultdict(lambda: defaultdict(list))
    for job in jobs:
        jobs_by_company[job["company"]][job["date"]].append(job)

    rows = []
    for company, days in jobs_by_company.items():
        rows.append([company])
        for date, entries in sorted(days.items()):
            rows.append([date])
            for job in entries:
                rows.append([
                    job['time'], job['name'], job['cid'],
                    job['type'], job['address'], f"WO {job['wo']}"
                ])
            rows.append([])
        rows.append([])

    # Write to Excel
    df = pd.DataFrame(rows)
    df.to_excel(filename, index=False, header=False)

    # Auto-adjust column widths
    wb = load_workbook(filename)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add padding

    wb.save(filename)

# === COOKIE & LOGIN FUNCTIONS ===
def save_cookies(driver, filename="cookies.pkl"):
    import pickle
    with open(filename, "wb") as f:
        pickle.dump(driver.get_cookies(), f)

def load_cookies(driver, filename="cookies.pkl"):
    import pickle

    if not os.path.exists(filename):
        return False

    try:
        with open(filename, "rb") as f:
            cookies = pickle.load(f)
            if not isinstance(cookies, list):
                print("❌ Cookie file is corrupted or not a list.")
                return False

        driver.get("http://inside.sockettelecom.com/")
        for cookie in cookies:
            if not isinstance(cookie, dict):
                print("❌ Skipping invalid cookie:", cookie)
                continue
            driver.add_cookie(cookie)

        driver.refresh()
        time.sleep(1)

        # Confirm login is valid
        if "login.php" in driver.current_url or "Username" in driver.page_source:
            print("🔐 Cookies invalid — still on login page.")
            return False

        clear_first_time_overlays(driver)
        return True

    except Exception as e:
        print(f"⚠️ Failed to load cookies: {e}")
        if os.path.exists(filename):
            os.remove(filename)
            print("🧹 Removed bad cookie file.")
        return False


def handle_login(driver):
    driver.get("http://inside.sockettelecom.com/")
    if not load_cookies(driver):
        print("🔐 Logging in manually...")
        try:
            perform_login(driver)
            time.sleep(2)
            save_cookies(driver)
        except RuntimeError as e:
            print("❌ Could not complete login. Check credentials or server.")
            raise e
    else:
        print("🍪 Logged in via cookies.")



def perform_login(driver):
    driver.get("http://inside.sockettelecom.com/system/login.php")
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.NAME, "username")))

    dismiss_alert(driver)

    driver.find_element(By.NAME, "username").send_keys(USERNAME)
    driver.find_element(By.NAME, "password").send_keys(PASSWORD)
    driver.find_element(By.ID, "login").click()

    dismiss_alert(driver)

    try:
        WebDriverWait(driver, 10).until(
            lambda d: "menu.php" in d.current_url or "Dashboard" in d.page_source
        )
    except TimeoutException:
        print("❌ Login failed or server unresponsive.")
        raise RuntimeError("Login failed")

    clear_first_time_overlays(driver)
    print("✅ Login complete and overlays cleared.")


if __name__ == "__main__":
    log_file = "calendarbuddy_crashlog.txt"
    try:
        options = Options()
        options.page_load_strategy = 'eager'
        if HEADLESS:
            options.add_argument("--headless=new")
        options.add_experimental_option("detach", True)

        root = TkinterDnD.Tk()
        app = CalendarBuddyGUI(root)
        root.mainloop()

    except Exception as e:
        error_message = f"❌ Crash occurred: {e}\n"
        trace = traceback.format_exc()
        print(error_message)
        print(trace)

        with open(log_file, "w", encoding="utf-8") as f:
            f.write("=== Calendar Buddy Crash Log ===\n")
            f.write(error_message)
            f.write(trace)

        print(f"📄 Error log written to: {log_file}")

    finally:
        try:
            if hasattr(app, "driver") and app.driver:
                app.driver.quit()
        except:
            pass
        input("✅ Done! Press Enter to close...")
